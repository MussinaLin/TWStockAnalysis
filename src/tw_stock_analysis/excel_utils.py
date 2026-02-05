from __future__ import annotations

import datetime as dt
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def load_history(
    path: Path,
) -> tuple[dict[str, pd.Series], dict[str, pd.Series]]:
    """Load historical close prices and volumes from Excel sheets.

    Returns:
        Tuple of (close_history, volume_history) where each is a dict
        mapping symbol to a pandas Series indexed by date.
    """
    if not path.exists():
        return {}, {}

    close_hist: dict[str, list[tuple[dt.date, float]]] = {}
    volume_hist: dict[str, list[tuple[dt.date, float]]] = {}
    xls = pd.ExcelFile(path)
    for sheet in xls.sheet_names:
        try:
            sheet_date = dt.date.fromisoformat(sheet)
        except ValueError:
            continue
        df = xls.parse(sheet)
        if "symbol" not in df.columns or "close" not in df.columns:
            continue
        for _, row in df.iterrows():
            symbol = str(row["symbol"]).strip()
            close = row["close"]
            if pd.notna(close):
                close_hist.setdefault(symbol, []).append((sheet_date, float(close)))
            vol = row.get("volume")
            if pd.notna(vol):
                volume_hist.setdefault(symbol, []).append((sheet_date, float(vol)))

    def _to_series(hist: dict[str, list[tuple[dt.date, float]]]) -> dict[str, pd.Series]:
        result: dict[str, pd.Series] = {}
        for symbol, rows in hist.items():
            rows = sorted(rows, key=lambda item: item[0])
            dates = [item[0] for item in rows]
            values = [item[1] for item in rows]
            result[symbol] = pd.Series(values, index=pd.to_datetime(dates))
        return result

    return _to_series(close_hist), _to_series(volume_hist)


def get_sheet_names(path: Path) -> set[str]:
    if not path.exists():
        return set()
    xls = pd.ExcelFile(path)
    return set(xls.sheet_names)


def write_daily_sheet(path: Path, date: dt.date, df: pd.DataFrame) -> None:
    sheet_name = date.isoformat()
    if path.exists():
        with pd.ExcelWriter(
            path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def write_market_closed_sheet(
    path: Path, date: dt.date, reason: str, details: str
) -> None:
    sheet_name = "market_closed"
    row = {"date": date.isoformat(), "reason": reason, "details": details}

    if path.exists():
        xls = pd.ExcelFile(path)
        if sheet_name in xls.sheet_names:
            df = xls.parse(sheet_name)
        else:
            df = pd.DataFrame(columns=["date", "reason", "details"])
    else:
        df = pd.DataFrame(columns=["date", "reason", "details"])

    if "date" not in df.columns:
        df["date"] = None
    if "reason" not in df.columns:
        df["reason"] = None
    if "details" not in df.columns:
        df["details"] = None

    df = df[df["date"].astype(str) != row["date"]]
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    df["_date_sort"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.sort_values("_date_sort").drop(columns="_date_sort")

    if path.exists():
        with pd.ExcelWriter(
            path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def remove_sheet(path: Path, sheet_name: str) -> None:
    if not path.exists():
        return
    workbook = load_workbook(path)
    if sheet_name not in workbook.sheetnames:
        return
    worksheet = workbook[sheet_name]
    workbook.remove(worksheet)
    workbook.save(path)
